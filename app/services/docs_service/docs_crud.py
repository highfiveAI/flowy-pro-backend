# app/services/docs_service/docs_crud.py

import os
from datetime import datetime
from typing import List, Optional
from uuid import UUID
import tempfile
import aiofiles
import aioboto3
import PyPDF2
import docx
import openpyxl
from botocore.exceptions import ClientError
from dotenv import load_dotenv
from fastapi import UploadFile, HTTPException
from sentence_transformers import SentenceTransformer
from openai import AsyncOpenAI
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy.future import select

from app.models.interdoc import Interdoc

# 환경 변수 로드
load_dotenv()

# DB 연결 설정 (비동기)
DATABASE_URL = os.getenv('CONNECTION_STRING').replace('postgresql://', 'postgresql+asyncpg://')
engine = create_async_engine(DATABASE_URL, echo=True)
AsyncSessionLocal = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

async def get_db():
    async with AsyncSessionLocal() as session:
        try:
            yield session
        finally:
            await session.close()

# S3 클라이언트 설정 (비동기)
session = aioboto3.Session(
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY'),
    aws_secret_access_key=os.getenv('AWS_SECRET_KEY'),
    region_name=os.getenv('AWS_REGION')
)

# OpenAI 클라이언트 설정 (비동기)
openai_client = AsyncOpenAI(api_key=os.getenv('OPENAI_API_KEY'))

# 문서 임베딩 모델 초기화
model = SentenceTransformer('sentence-transformers/distiluse-base-multilingual-cased-v2')

async def read_file_content(file: UploadFile) -> str:
    """파일 형식에 따라 내용을 읽는 함수"""
    content = ""
    file_ext = file.filename.lower().split('.')[-1]
    
    await file.seek(0) # 중요: 파일 포인터 초기화
    
    try:
        if file_ext == 'txt':
            content = (await file.read()).decode('utf-8')
        
        elif file_ext == 'pdf':
            # 임시 파일로 저장 후 처리
            async with aiofiles.tempfile.NamedTemporaryFile(delete=False) as temp_file:
                await temp_file.write(await file.read())
                temp_path = temp_file.name

            try:
                pdf_reader = PyPDF2.PdfReader(temp_path)
                for page in pdf_reader.pages:
                    content += page.extract_text() + "\n"
            finally:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
        
        elif file_ext in ['doc', 'docx']:
            # 임시 파일로 저장 후 처리
            async with aiofiles.tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as temp_file:
                await temp_file.write(await file.read())
                temp_path = temp_file.name

            try:
                doc = docx.Document(temp_path)
                content = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            finally:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            
        elif file_ext in ['xlsx', 'xls']:
            # 임시 파일로 저장 후 처리
            async with aiofiles.tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as temp_file:
                await temp_file.write(await file.read())
                temp_path = temp_file.name
            
            try:
                workbook = openpyxl.load_workbook(temp_path, data_only=True)
                sheets_content = []
                
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    sheet_content = []
                    
                    # 시트 이름 추가
                    sheet_content.append(f"[시트: {sheet}]")
                    
                    # 각 행의 데이터를 읽음
                    for row in worksheet.iter_rows():
                        row_values = [str(cell.value) if cell.value is not None else '' for cell in row]
                        if any(row_values):  # 빈 행 제외
                            sheet_content.append(' | '.join(row_values))
                    
                    sheets_content.append('\n'.join(sheet_content))
                
                content = '\n\n'.join(sheets_content)
                
            finally:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
        
        else:
            raise HTTPException(
                status_code=400,
                detail="지원하지 않는 파일 형식입니다. (지원 형식: txt, pdf, doc, docx, xlsx, xls)"
            )
            
        if not content.strip():
            raise HTTPException(
                status_code=400,
                detail="파일에서 텍스트를 추출할 수 없습니다."
            )
            
        return content
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"파일 읽기 실패: {str(e)}"
        )

async def extract_text_from_file(file: UploadFile) -> str:
    """LLM을 이용해 파일 내용을 요약하는 함수"""
    try:
        await file.seek(0)
        content = await read_file_content(file)
        
        response = await openai_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "주어진 문서의 전체적인 구조를 파악하여, 1~2문장 요약을 작성해주세요. 예를 들어, '목차, 목표, 일정, 예산 등으로 구성된 프로젝트 기획안'과 같이, 문서의 구성 요소를 중심으로 문서 종류와 연결되는 자연스러운 문장으로 만들어 주세요. 문서의 구체적 내용보다는 형식적·조직적 구성을 중심으로 정리해 주세요."},
                {"role": "user", "content": content}
            ],
            max_tokens=300,
            temperature=0.3
        )
        
        summary = response.choices[0].message.content.strip()
        return summary
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"문서 요약 실패: {str(e)}"
        )

async def create_document(
    db: AsyncSession,
    file: UploadFile,
    doc_type: str,
    update_user_id: UUID
) -> Interdoc:
    """문서 생성 함수"""
    try:
        content = await extract_text_from_file(file)
        embedding = model.encode(content)
        
        s3_path = f"documents/{file.filename}"
        
        await file.seek(0)
        async with session.client('s3') as s3:
            await s3.upload_fileobj(
                file.file,
                os.getenv('AWS_BUCKET_NAME'),
                s3_path
            )
        
        doc = Interdoc(
            interdocs_type_name=doc_type,
            interdocs_filename=file.filename,
            interdocs_contents=content[:255],
            interdocs_vector=embedding,
            interdocs_path=s3_path,
            interdocs_uploaded_date=datetime.now(),
            interdocs_update_user_id=update_user_id
        )
        
        db.add(doc)
        await db.commit()
        await db.refresh(doc)
        
        return doc
        
    except Exception as e:
        if 's3_path' in locals():
            try:
                async with session.client('s3') as s3:
                    await s3.delete_object(
                        Bucket=os.getenv('AWS_BUCKET_NAME'),
                        Key=s3_path
                    )
            except:
                pass
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

async def update_document(
    db: AsyncSession,
    doc_id: UUID,
    file: UploadFile,
    update_user_id: UUID
) -> Interdoc:
    """문서 수정 함수"""
    try:
        query = select(Interdoc).filter(Interdoc.interdocs_id == doc_id)
        result = await db.execute(query)
        doc = result.scalar_one_or_none()
        
        if not doc:
            raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
            
        content = await extract_text_from_file(file)
        embedding = model.encode(content)
        
        old_s3_path = doc.interdocs_path
        new_s3_path = f"documents/{file.filename}"
        
        await file.seek(0)
        async with session.client('s3') as s3:
            await s3.upload_fileobj(
                file.file,
                os.getenv('AWS_BUCKET_NAME'),
                new_s3_path
            )
            
            await s3.delete_object(
                Bucket=os.getenv('AWS_BUCKET_NAME'),
                Key=old_s3_path
            )
        
        doc.interdocs_filename = file.filename
        doc.interdocs_contents = content[:255]
        doc.interdocs_vector = embedding
        doc.interdocs_path = new_s3_path
        doc.interdocs_updated_date = datetime.now()
        doc.interdocs_update_user_id = update_user_id
        
        await db.commit()
        await db.refresh(doc)
        
        return doc
        
    except Exception as e:
        if 'new_s3_path' in locals():
            try:
                async with session.client('s3') as s3:
                    await s3.delete_object(
                        Bucket=os.getenv('AWS_BUCKET_NAME'),
                        Key=new_s3_path
                    )
            except:
                pass
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

async def get_documents(
    db: AsyncSession,
    skip: int = 0,
    limit: int = 10
) -> List[Interdoc]:
    """문서 목록 조회 함수"""
    print("get_documents 함수 시작")
    try:
        print(f"DB 세션 객체: {db}")
        print(f"Interdoc 모델: {Interdoc}")
        
        query = select(Interdoc).offset(skip).limit(limit)
        result = await db.execute(query)
        documents = result.scalars().all()
        
        print(f"조회된 문서 수: {len(documents)}")
        return documents
    except Exception as e:
        print(f"!!! get_documents 에러 발생: {e}")
        raise HTTPException(status_code=500, detail=f"문서 목록 조회 실패: {str(e)}")

async def get_document(
    db: AsyncSession,
    doc_id: UUID
) -> Optional[Interdoc]:
    """단일 문서 조회 함수"""
    try:
        query = select(Interdoc).filter(Interdoc.interdocs_id == doc_id)
        result = await db.execute(query)
        return result.scalar_one_or_none()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"단일 문서 조회 실패: {str(e)}")

async def delete_document(
    db: AsyncSession,
    doc_id: UUID
) -> bool:
    """문서 삭제 함수"""
    try:
        query = select(Interdoc).filter(Interdoc.interdocs_id == doc_id)
        result = await db.execute(query)
        doc = result.scalar_one_or_none()
        
        if not doc:
            raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다")
            
        async with session.client('s3') as s3:
            await s3.delete_object(
                Bucket=os.getenv('AWS_BUCKET_NAME'),
                Key=doc.interdocs_path
            )
        
        await db.delete(doc)
        await db.commit()
        
        return True
        
    except ClientError as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"S3 삭제 실패: {str(e)}")
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))